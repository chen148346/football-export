"""
excel_exporter.py - Excel 导出模块
====================================
将比赛数据导出为包含7个Sheet的Excel文件。

7个Sheet：
1. match_overview   - 比赛总览（每行一场比赛）
2. tech_stats       - 技术统计（动态列，每行一场比赛）
3. events           - 重要事件（每行一个事件）
4. player_stats     - 球员统计（每行一名球员，动态技术指标列）
5. near_matches     - 近期战绩（每行一场近期比赛）
6. vs_matches       - 交锋历史（每行一场交锋记录）
7. goal_distribution - 进失球分布（每行一场比赛×count_type×球队）

技术要点：
- openpyxl 生成 Excel
- tech_stats 和 player_stats 使用动态列（扫描所有比赛取并集）
- 表头加粗、冻结首行、自动列宽
- 空值统一为空字符串
"""

import os
import datetime
from typing import List, Dict, Any, Optional
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import config
from . import json_parser
from .db_reader import MatchRecord


def _beijing_now() -> datetime.datetime:
    """返回当前北京时间（UTC+8）"""
    utc_now = datetime.datetime.utcnow()
    return utc_now + datetime.timedelta(hours=config.BEIJING_TZ_OFFSET_HOURS)


# ============================================================================
# 样式定义
# ============================================================================

# 表头样式
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 数据样式
DATA_FONT = Font(name="Microsoft YaHei", size=10)
DATA_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)

# 边框
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _apply_header_style(ws, row_idx: int, col_count: int):
    """应用表头样式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _apply_data_style(ws, start_row: int, end_row: int, col_count: int):
    """应用数据样式"""
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER


def _auto_column_width(ws, col_count: int, max_width: int = 40):
    """自动调整列宽（基于表头和前若干行数据）"""
    for col in range(1, col_count + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        # 检查表头和前100行数据
        for row in range(1, min(ws.max_row, 100) + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                # 中文字符算2个宽度
                length = sum(2 if ord(c) > 127 else 1 for c in str(val))
                if length > max_len:
                    max_len = length
        width = min(max_len + 2, max_width)
        ws.column_dimensions[col_letter].width = max(width, 8)


def _write_sheet(ws, headers: List[str], rows: List[Dict[str, Any]]):
    """
    通用写入函数：写入表头和数据行。
    
    headers: 列名列表
    rows: 数据行列表，每个元素是字典，键名与 headers 对应
    """
    # 写表头
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _apply_header_style(ws, 1, len(headers))
    
    # 写数据
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row_data.get(header)
            # 确保值是基本类型（openpyxl 不支持复杂对象）
            # None 和缺失的key都转为空字符串
            if val is None:
                val = ""
            elif isinstance(val, (dict, list)):
                val = str(val)
            ws.cell(row=row_idx, column=col_idx, value=val)
    
    # 应用数据样式
    if rows:
        _apply_data_style(ws, 2, len(rows) + 1, len(headers))
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 自动列宽
    _auto_column_width(ws, len(headers))


# ============================================================================
# Sheet 1: match_overview - 比赛总览
# ============================================================================

def build_match_overview(matches: List[MatchRecord]) -> tuple:
    """
    构建比赛总览数据。
    
    每行一场比赛，包含基本信息、比分、排名、阵型、角球、红黄牌等。
    字段来源：matches表（数据库字段）+ shijian_json（阵型/角球）+ analysis_json（裁判）。
    
    V1.5变更说明（比分字段分离）：
    - 原 home_score/away_score 拆分为：
      - current_home_score/current_away_score: 当前时点比分（特征，赛中可用）
      - label_home_score/label_away_score: 完场比分（标签，仅完场快照有值）
    - 消除"同一字段既作特征又作标签"的语义歧义，防止ML数据泄露
    - 完场快照：current=label=完场比分
    - 非完场快照：current=当前时点比分，label=完场比分（需关联完场快照）
    """
    headers = [
        "match_id", "sclass_name", "match_date", "match_time",
        "home_team", "away_team",
        "home_rank", "away_rank",
        # V1.5: 比分字段分离
        "current_home_score", "current_away_score",   # 当前时点比分（特征）
        "label_home_score", "label_away_score",        # 完场比分（标签）
        "home_half_score", "away_half_score",
        "home_red", "away_red",
        "home_yellow", "away_yellow",
        "state_code", "state_text", "elapsed_min",
        "weather", "round_info", "is_neutrality",
        "home_formation", "away_formation",
        "home_corner", "away_corner",
        "home_half_corner", "away_half_corner",
        "referee_name",
    ]
    
    rows = []
    for m in matches:
        # 从 shijian_json 提取阵型和角球信息
        home_formation = ""
        away_formation = ""
        home_corner = ""
        away_corner = ""
        home_half_corner = ""
        away_half_corner = ""
        referee_name = ""
        
        if m.shijian_json:
            lineup = json_parser.extract_lineup(m.shijian_json)
            home_formation = lineup["home_formation"]
            away_formation = lineup["away_formation"]
            
            corner = json_parser.extract_corner_events(m.shijian_json)
            home_corner = corner["home_corner"]
            away_corner = corner["away_corner"]
            home_half_corner = corner["home_half_corner"]
            away_half_corner = corner["away_half_corner"]
        
        if m.analysis_json:
            ref = json_parser.extract_referee(m.analysis_json)
            referee_name = ref["referee_name"]
        
        # 格式化比赛时间
        match_time_dt = json_parser.parse_match_time_14(m.match_time)
        match_time_str = json_parser.format_datetime(match_time_dt)
        
        # V1.5: 比分字段分离逻辑
        # current_score: 当前时点比分（来自快照的home_score/away_score）
        # label_score: 完场比分（完场快照=current；非完场快照需关联完场快照，暂用current填充）
        current_home = m.home_score
        current_away = m.away_score
        
        # 判断是否为完场快照
        if m.snapshot_type == "fulltime" or m.state_code == -1:
            # 完场快照：label = current
            label_home = m.home_score
            label_away = m.away_score
        else:
            # 非完场快照：label需要关联完场快照获取
            # 当前实现中，非完场快照的label通过关联查询填充
            # 如果未关联，label留空（由调用方通过关联完场快照补充）
            label_home = ""
            label_away = ""
        
        rows.append({
            "match_id": m.match_id,
            "sclass_name": m.sclass_name,
            "match_date": m.match_date,
            "match_time": match_time_str,
            "home_team": m.home_team,
            "away_team": m.away_team,
            "home_rank": m.home_rank,
            "away_rank": m.away_rank,
            # V1.5: 比分字段分离
            "current_home_score": current_home,
            "current_away_score": current_away,
            "label_home_score": label_home,
            "label_away_score": label_away,
            "home_half_score": m.home_half_score,
            "away_half_score": m.away_half_score,
            "home_red": m.home_red,
            "away_red": m.away_red,
            "home_yellow": m.home_yellow,
            "away_yellow": m.away_yellow,
            "state_code": m.state_code,
            "state_text": m.state_text,
            "elapsed_min": m.elapsed_min,
            "weather": m.weather,
            "round_info": m.round_info,
            "is_neutrality": m.is_neutrality,
            "home_formation": home_formation,
            "away_formation": away_formation,
            "home_corner": home_corner,
            "away_corner": away_corner,
            "home_half_corner": home_half_corner,
            "away_half_corner": away_half_corner,
            "referee_name": referee_name,
        })
    
    return headers, rows


# ============================================================================
# Sheet 2: tech_stats - 技术统计（动态列）
# ============================================================================

def build_tech_stats(matches: List[MatchRecord]) -> tuple:
    """
    构建技术统计数据。
    
    动态列：扫描所有比赛的技术统计项，取 kind 并集。
    每个统计项生成两列：home_{kind} 和 away_{kind}。
    """
    # 第一遍：收集所有 kind（保持出现顺序）
    all_kinds = OrderedDict()  # kind -> name
    for m in matches:
        if not m.shijian_json:
            continue
        tech = json_parser.extract_tech_stats(m.shijian_json, m.match_id)
        for kind, info in tech["stats"].items():
            if kind not in all_kinds:
                all_kinds[kind] = info["name"]
    
    # 构建表头
    headers = ["match_id", "sclass_name", "home_team", "away_team"]
    for kind, name in all_kinds.items():
        # 列名格式：home_{kind}_{name} 和 away_{kind}_{name}
        label = f"{kind}_{all_kinds[kind]}" if all_kinds[kind] else kind
        headers.append(f"home_{label}")
        headers.append(f"away_{label}")
    
    # 第二遍：构建数据行
    rows = []
    for m in matches:
        row = {
            "match_id": m.match_id,
            "sclass_name": m.sclass_name,
            "home_team": m.home_team,
            "away_team": m.away_team,
        }
        
        if m.shijian_json:
            tech = json_parser.extract_tech_stats(m.shijian_json, m.match_id)
            for kind, info in tech["stats"].items():
                label = f"{kind}_{all_kinds[kind]}" if all_kinds[kind] else kind
                row[f"home_{label}"] = info["home"]
                row[f"away_{label}"] = info["away"]
        
        rows.append(row)
    
    return headers, rows


# ============================================================================
# Sheet 3: events - 重要事件
# ============================================================================

def build_events(matches: List[MatchRecord]) -> tuple:
    """构建重要事件数据（每行一个事件）"""
    headers = [
        "match_id", "event_index", "time", "side", "process",
        "event_type", "player_name", "player_id",
        "related_player", "home_score", "away_score", "extra_info",
    ]
    
    rows = []
    for m in matches:
        if not m.shijian_json:
            continue
        events = json_parser.extract_events(m.shijian_json, m.match_id)
        rows.extend(events)
    
    return headers, rows


# ============================================================================
# Sheet 4: player_stats - 球员统计（动态技术指标列）
# ============================================================================

def build_player_stats(matches: List[MatchRecord]) -> tuple:
    """
    构建球员统计数据。
    
    动态列：扫描所有球员的 techInfos，取 infoKind 并集。
    """
    # 第一遍：收集所有球员数据和所有 tech infoKind
    all_players = []
    all_tech_kinds = set()
    
    for m in matches:
        if not m.shijian_json:
            continue
        players = json_parser.extract_player_stats(m.shijian_json, m.match_id)
        for p in players:
            all_players.append(p)
            for key in p.keys():
                if key.startswith("tech_"):
                    all_tech_kinds.add(key)
    
    # 固定列
    fixed_headers = [
        "match_id", "team_side", "team_name", "formation",
        "player_id", "player_name", "player_num", "is_best",
    ]
    
    # 动态列（排序后）
    dynamic_headers = sorted(all_tech_kinds)
    
    headers = fixed_headers + dynamic_headers
    
    # 构建数据行
    rows = []
    for p in all_players:
        row = {}
        for h in headers:
            row[h] = p.get(h, "")
        rows.append(row)
    
    return headers, rows


# ============================================================================
# Sheet 5: near_matches - 近期战绩
# ============================================================================

def build_near_matches(matches: List[MatchRecord]) -> tuple:
    """构建近期战绩数据（每行一场近期比赛）"""
    headers = [
        "match_id", "team_side", "team_name",
        "near_match_id", "league_name", "league_name_full",
        "match_time", "match_time_ts",
        "home_team_name", "away_team_name",
        "home_score", "away_score",
        "home_half_score", "away_half_score",
        "home_corner", "away_corner",
        "home_half_corner", "away_half_corner",
        "home_yellow", "away_yellow",
        "home_red", "away_red",
        "home_is_first_goal", "away_is_first_goal",
        "home_shot_on_target", "away_shot_on_target",
        "is_neutrality",
    ]
    
    rows = []
    for m in matches:
        if not m.analysis_json:
            continue
        near = json_parser.extract_near_matches(m.analysis_json, m.match_id)
        rows.extend(near)
    
    return headers, rows


# ============================================================================
# Sheet 6: vs_matches - 交锋历史
# ============================================================================

def build_vs_matches(matches: List[MatchRecord]) -> tuple:
    """构建交锋历史数据（每行一场交锋记录）"""
    headers = [
        "match_id",
        "near_match_id", "league_name", "league_name_full",
        "match_time", "match_time_ts",
        "home_team_name", "away_team_name",
        "home_score", "away_score",
        "home_half_score", "away_half_score",
        "home_corner", "away_corner",
        "home_half_corner", "away_half_corner",
        "home_yellow", "away_yellow",
        "home_red", "away_red",
        "home_is_first_goal", "away_is_first_goal",
        "home_shot_on_target", "away_shot_on_target",
        "is_neutrality",
    ]
    
    rows = []
    for m in matches:
        if not m.analysis_json:
            continue
        vs = json_parser.extract_vs_matches(m.analysis_json, m.match_id)
        rows.extend(vs)
    
    return headers, rows


# ============================================================================
# Sheet 7: goal_distribution - 进失球分布
# ============================================================================

def build_goal_distribution(matches: List[MatchRecord]) -> tuple:
    """
    构建进失球分布数据。
    
    每行一场比赛×count_type×球队，共最多4行/场。
    6个时间段，每个时间段有 JQ（进球率）和 SQ（失球率）。
    """
    headers = [
        "match_id", "count_type", "team_side",
    ]
    # 6个时间段的 JQ 和 SQ
    for period in config.GOAL_DIST_TIME_PERIODS:
        headers.append(f"{period}_JQ")
        headers.append(f"{period}_SQ")
    
    rows = []
    for m in matches:
        if not m.shijian_json:
            continue
        gd = json_parser.extract_goal_distribution(m.shijian_json, m.match_id)
        rows.extend(gd)
    
    return headers, rows


# ============================================================================
# 主导出函数
# ============================================================================

def generate_filename(
    sclass_names: List[str],
    start_date: str,
    end_date: str,
) -> str:
    """
    生成导出文件名。
    
    格式：training_{联赛}_{开始日期}_{结束日期}_{时间戳}.xlsx
    多联赛时用 "multi" 代替。
    """
    if sclass_names and len(sclass_names) == 1:
        league = sclass_names[0]
    elif sclass_names and len(sclass_names) > 1:
        league = "multi"
    else:
        league = "all"
    
    # 清理联赛名中的特殊字符
    league = league.replace("/", "_").replace("\\", "_").replace(" ", "")
    
    start = start_date or "all"
    end = end_date or "all"
    timestamp = _beijing_now().strftime("%Y%m%d%H%M%S")
    
    return config.EXPORT_FILENAME_TEMPLATE.format(
        league=league,
        start_date=start,
        end_date=end,
        timestamp=timestamp,
    )


# ============================================================================
# V1.4 新增 Sheet 8: first_half_tech_stats - 上半场技术统计（动态列）
# ============================================================================

def build_first_half_tech_stats(matches: List[MatchRecord]) -> tuple:
    """
    V1.4新增 - 构建上半场技术统计数据（动态列）。
    
    数据源：shijian_json.techStat.firstHalfList
    注意：firstHalfList长度可能与itemList不同，按kind动态匹配。
    百分比字段已小数化。
    """
    return _build_half_tech_stats(matches, "first")


def build_second_half_tech_stats(matches: List[MatchRecord]) -> tuple:
    """
    V1.4新增 - 构建下半场技术统计数据（动态列）。
    
    数据源：shijian_json.techStat.secondHalfList
    """
    return _build_half_tech_stats(matches, "second")


def _build_half_tech_stats(matches: List[MatchRecord], half: str) -> tuple:
    """半场技术统计通用构建函数"""
    # 第一遍：收集所有kind
    all_kinds = OrderedDict()
    for m in matches:
        if not m.shijian_json:
            continue
        tech = json_parser.extract_half_tech_stats(m.shijian_json, m.match_id, half)
        for kind, info in tech["stats"].items():
            if kind not in all_kinds:
                all_kinds[kind] = info["name"]
    
    # 构建表头
    headers = ["match_id", "sclass_name", "home_team", "away_team"]
    for kind, name in all_kinds.items():
        label = f"{kind}_{all_kinds[kind]}" if all_kinds[kind] else kind
        headers.append(f"home_{label}")
        headers.append(f"away_{label}")
    
    # 构建数据行
    rows = []
    for m in matches:
        row = {
            "match_id": m.match_id,
            "sclass_name": m.sclass_name,
            "home_team": m.home_team,
            "away_team": m.away_team,
        }
        if m.shijian_json:
            tech = json_parser.extract_half_tech_stats(m.shijian_json, m.match_id, half)
            for kind, info in tech["stats"].items():
                label = f"{kind}_{all_kinds[kind]}" if all_kinds[kind] else kind
                row[f"home_{label}"] = info["home"]
                row[f"away_{label}"] = info["away"]
        rows.append(row)
    
    return headers, rows


# ============================================================================
# V1.4 新增 Sheet 10: league_rank_stats - 联赛排名统计（动态列）
# ============================================================================

def build_league_rank_stats(matches: List[MatchRecord]) -> tuple:
    """
    V1.4新增 - 构建联赛排名统计数据（动态列）。
    
    数据源：analysis_json.curLeagueStat.itemList
    注意：curLeagueStat可能为None（杯赛），需判空。
    homeValue/awayValue是字符串，已根据kind转int或float。
    """
    # 第一遍：收集所有kind
    all_kinds = OrderedDict()
    for m in matches:
        if not m.analysis_json:
            continue
        lrs = json_parser.extract_league_rank_stats(m.analysis_json, m.match_id)
        for kind, info in lrs["stats"].items():
            if kind not in all_kinds:
                all_kinds[kind] = info["name"]
    
    # 构建表头
    headers = ["match_id", "sclass_name", "home_team", "away_team"]
    for kind, name in all_kinds.items():
        label = f"{kind}_{all_kinds[kind]}" if all_kinds[kind] else kind
        headers.append(f"home_{label}")
        headers.append(f"away_{label}")
    
    # 构建数据行
    rows = []
    for m in matches:
        row = {
            "match_id": m.match_id,
            "sclass_name": m.sclass_name,
            "home_team": m.home_team,
            "away_team": m.away_team,
        }
        if m.analysis_json:
            lrs = json_parser.extract_league_rank_stats(m.analysis_json, m.match_id)
            for kind, info in lrs["stats"].items():
                label = f"{kind}_{all_kinds[kind]}" if all_kinds[kind] else kind
                row[f"home_{label}"] = info["home"]
                row[f"away_{label}"] = info["away"]
        rows.append(row)
    
    return headers, rows


# ============================================================================
# V1.4 新增 Sheet 11: half_full_stats - 近两赛季半场/全场统计
# ============================================================================

def build_half_full_stats(matches: List[MatchRecord]) -> tuple:
    """
    V1.4新增 - 构建近两赛季半场/全场统计数据。
    
    数据源：shijian_json.allhalf.list
    每行一个HA类型，9种组合：HA33, HA31, HA30, HA13, HA11, HA10, HA03, HA01, HA00
    """
    headers = [
        "match_id", "sclass_name", "home_team", "away_team",
        "ha_type", "ha_desc",
        "home_half", "home_all", "away_half", "away_all",
    ]
    
    rows = []
    for m in matches:
        if not m.shijian_json:
            continue
        hf_stats = json_parser.extract_half_full_stats(m.shijian_json, m.match_id)
        for item in hf_stats:
            rows.append({
                "match_id": m.match_id,
                "sclass_name": m.sclass_name,
                "home_team": m.home_team,
                "away_team": m.away_team,
                "ha_type": item["ha_type"],
                "ha_desc": item["ha_desc"],
                "home_half": item["home_half"],
                "home_all": item["home_all"],
                "away_half": item["away_half"],
                "away_all": item["away_all"],
            })
    
    return headers, rows


# ============================================================================
# V1.4 新增 Sheet 12: detailed_events - 详细事件
# ============================================================================

def build_detailed_events(matches: List[MatchRecord]) -> tuple:
    """
    V1.4新增 - 构建详细事件数据。
    
    数据源：shijian_json.eventTxt.EventTxtLives
    ⚠️ 已反转：从1'开赛到完场正序排列
    Context已剔除HTML标签
    """
    headers = [
        "match_id", "event_index", "time_txt",
        "happen_time", "match_state", "injure_time",
        "kind", "context",
    ]
    
    rows = []
    for m in matches:
        if not m.shijian_json:
            continue
        events = json_parser.extract_detailed_events(m.shijian_json, m.match_id)
        for ev in events:
            rows.append({
                "match_id": ev["match_id"],
                "event_index": ev["event_index"],
                "time_txt": ev["time_txt"],
                "happen_time": ev["happen_time"],
                "match_state": ev["match_state"],
                "injure_time": ev["injure_time"],
                "kind": ev["kind"],
                "context": ev["context"],
            })
    
    return headers, rows


def generate_filename_v12(
    sclass_names: List[str],
    start_date: str,
    end_date: str,
    sheets_to_export: List[str] = None,
    part_index: int = None,
) -> str:
    """
    V1.2新增 - 生成符合新命名规则的文件名。
    
    格式：{导出日期}_{赛事名称}_{开始日期}_{结束日期}_{后缀}.xlsx
    - 导出日期：yymmdd（如260708）
    - 赛事名称：单赛事用该名称，多赛事用/连接（文件名中替换为-）
    - 后缀：全选为all，部分选择用Sheet首字母拼接（如metv）
    - 分片时追加 _01, _02 等
    
    参数：
        sclass_names: 联赛名称列表
        start_date: 开始日期
        end_date: 结束日期
        sheets_to_export: 选中的Sheet标识列表（None=全选）
        part_index: 分片序号（None=不分片）
    """
    # 导出日期：yymmdd
    export_date = _beijing_now().strftime("%y%m%d")
    
    # 赛事名称
    if not sclass_names:
        league = "all"
    elif len(sclass_names) == 1:
        league = sclass_names[0]
    else:
        league = "/".join(sclass_names)
    
    # 安全处理文件名中的非法字符
    for illegal, safe in config.ILLEGAL_CHAR_REPLACE.items():
        league = league.replace(illegal, safe)
    
    # 日期
    start = start_date or "all"
    end = end_date or "all"
    
    # 后缀：全选为all，部分选择用首字母拼接
    if sheets_to_export is None or len(sheets_to_export) == 0:
        suffix = "all"
    else:
        # 检查是否全选
        all_ids = set(config.ALL_SHEET_IDS)
        selected = set(sheets_to_export)
        if all_ids.issubset(selected) or selected == all_ids:
            suffix = "all"
        else:
            # 按固定顺序拼接首字母
            letters = []
            for sid in config.ALL_SHEET_IDS:
                if sid in selected:
                    letters.append(config.SHEET_SUFFIX_LETTER.get(sid, ""))
            suffix = "".join(letters) if letters else "all"
    
    # 构建文件名
    filename = config.V12_FILENAME_TEMPLATE.format(
        export_date=export_date,
        league=league,
        start_date=start,
        end_date=end,
        suffix=suffix,
    )
    
    # 分片序号
    if part_index is not None:
        # 在.xlsx前插入_01
        filename = filename.replace(".xlsx", f"_{part_index:02d}.xlsx")
    
    return filename


def generate_filename_v15(
    sclass_names: List[str],
    start_date: str,
    end_date: str,
    sheets_to_export: List[str] = None,
    snapshot_type: str = "fulltime",
    file_role: str = None,
    part_index: int = None,
) -> str:
    """
    V1.5新增 - 生成符合V1.5新命名规则的文件名。
    
    格式：{时间戳}_{联赛}_{开始日期}_{结束日期}_{类型}_{字段选择}.xlsx
    - 时间戳：yymmddhhmm（精确到分钟，如2607141530）
    - 联赛：单联赛用中文名，多联赛统一用Multi
    - 开始/结束日期：yymmdd格式（如260710）
    - 类型：FT(完场)/HT(半场)/PR(进行中/自定义分钟区间)
    - 字段选择：all(12个全选) / part(没有全选12个就是part)
    
    参数：
        sclass_names: 联赛名称列表
        start_date: 开始日期 YYYY-MM-DD
        end_date: 结束日期 YYYY-MM-DD
        sheets_to_export: 选中的Sheet标识列表（None=全选）
        snapshot_type: 快照类型 fulltime/halftime/custom
        file_role: 文件角色 prediction/label（用于非完场导出的关联完场文件）
        part_index: 分片序号（None=不分片）
    
    V1.5变更说明：
    - 废弃V1.2的首字母拼接后缀，改为all/part二值
    - 时间戳精确到分钟（yymmddhhmm），避免同一天多次导出混淆
    - 新增快照类型标识，解决非完场/完场文件名冲突
    - 日期格式统一为yymmdd
    """
    # 时间戳：yymmddhhmm（精确到分钟）
    timestamp = _beijing_now().strftime("%y%m%d%H%M")
    
    # 联赛名称：单联赛用中文名，多联赛统一用Multi
    if not sclass_names:
        league = "all"
    elif len(sclass_names) == 1:
        league = sclass_names[0]
    else:
        league = "Multi"
    
    # 安全处理文件名中的非法字符
    for illegal, safe in config.ILLEGAL_CHAR_REPLACE.items():
        league = league.replace(illegal, safe)
    
    # 日期：yymmdd格式
    def format_date_yymmdd(date_str):
        if not date_str:
            return "all"
        try:
            # 支持 YYYY-MM-DD 格式
            dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
            return dt.strftime("%y%m%d")
        except (ValueError, TypeError):
            pass
        try:
            # 支持 ISO datetime 格式: YYYY-MM-DDTHH:MM
            dt = datetime.datetime.strptime(date_str.split("T")[0], "%Y-%m-%d")
            return dt.strftime("%y%m%d")
        except (ValueError, TypeError, IndexError):
            pass
        # 最后兜底：提取数字部分
        digits = "".join(ch for ch in date_str if ch.isdigit())
        return digits[-6:] if len(digits) >= 6 else "all"
    
    start = format_date_yymmdd(start_date)
    end = format_date_yymmdd(end_date)
    
    # 类型：FT/HT/PR
    snap_type = config.V15_SNAPSHOT_TYPE_MAP.get(snapshot_type, "FT")
    
    # 字段选择：all(12个全选) / part(没有全选12个就是part)
    all_12_ids = set(config.ALL_SHEET_IDS_V15)
    if sheets_to_export is None:
        field_select = "all"
    else:
        selected = set(sheets_to_export)
        # 检查是否12个全选
        if all_12_ids.issubset(selected):
            field_select = "all"
        else:
            field_select = "part"
    
    # 构建文件名
    filename = config.V15_FILENAME_TEMPLATE.format(
        timestamp=timestamp,
        league=league,
        start_date=start,
        end_date=end,
        snapshot_type=snap_type,
        field_select=field_select,
    )
    
    # 文件角色后缀（prediction/label）
    if file_role:
        filename = filename.replace(".xlsx", f"_{file_role}.xlsx")
    
    # 分片序号
    if part_index is not None:
        filename = filename.replace(".xlsx", f"_{part_index:02d}.xlsx")
    
    return filename


def export_to_excel(
    matches: List[MatchRecord],
    output_path: str = None,
    sclass_names: List[str] = None,
    start_date: str = None,
    end_date: str = None,
    sheets_to_export: List[str] = None,
) -> str:
    """
    将比赛数据导出为 Excel 文件。
    
    参数：
        matches: MatchRecord 列表
        output_path: 输出文件路径，None 则自动生成
        sclass_names: 联赛名称列表（用于文件名）
        start_date: 开始日期（用于文件名）
        end_date: 结束日期（用于文件名）
        sheets_to_export: 要导出的Sheet标识列表，None表示全选。
            V1.0的7个: match, tech, events, player, near, vs, dist
            V1.4的5个: first_half, second_half, league_rank, half_full, detail_evt
            match为必选项（即使未传入也会自动添加）
    
    返回：
        生成的 Excel 文件路径
    
    V1.4变更说明：
        - 支持V1.4新增的5个Sheet（first_half, second_half, league_rank, half_full, detail_evt）
        - 使用config.SHEET_ID_MAP_V14（12个Sheet的完整映射）
        - 默认全选时仅导出V1.0的7个Sheet（保持向后兼容）
        - V1.4的5个Sheet需显式选择才导出
    """
    # V1.4: 使用完整的12个Sheet映射
    SHEET_ID_MAP = config.SHEET_ID_MAP_V14
    
    # V1.4: 确定要导出的Sheet列表
    if sheets_to_export is None:
        # 默认全选V1.0的7个Sheet（保持向后兼容）
        selected_sheet_names = [
            "match_overview", "tech_stats", "events",
            "player_stats", "near_matches", "vs_matches",
            "goal_distribution",
        ]
    else:
        # 根据用户选择构建Sheet列表
        selected_sheet_names = []
        for sid in sheets_to_export:
            if sid in SHEET_ID_MAP:
                sname = SHEET_ID_MAP[sid]
                if sname not in selected_sheet_names:
                    selected_sheet_names.append(sname)
        # match_overview 必选
        if "match_overview" not in selected_sheet_names:
            selected_sheet_names.insert(0, "match_overview")
    
    if output_path is None:
        filename = generate_filename(sclass_names, start_date, end_date)
        output_path = os.path.join(config.OUTPUT_DIR, filename)
    
    # 确保输出目录存在
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    
    wb = Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)
    
    # V1.4: 完整的12个Sheet构建器
    all_builders = [
        # V1.0的7个
        ("match_overview", build_match_overview),
        ("tech_stats", build_tech_stats),
        ("events", build_events),
        ("player_stats", build_player_stats),
        ("near_matches", build_near_matches),
        ("vs_matches", build_vs_matches),
        ("goal_distribution", build_goal_distribution),
        # V1.4的5个
        ("first_half_tech_stats", build_first_half_tech_stats),
        ("second_half_tech_stats", build_second_half_tech_stats),
        ("league_rank_stats", build_league_rank_stats),
        ("half_full_stats", build_half_full_stats),
        ("detailed_events", build_detailed_events),
    ]
    
    sheet_stats = {}
    for sheet_name, builder in all_builders:
        # 仅构建选中的Sheet
        if sheet_name not in selected_sheet_names:
            continue
        headers, rows = builder(matches)
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, headers, rows)
        sheet_stats[sheet_name] = len(rows)
        print(f"  [Sheet] {sheet_name}: {len(rows)} 行, {len(headers)} 列")
    
    wb.save(output_path)
    print(f"\n导出完成: {output_path}")
    print(f"文件大小: {os.path.getsize(output_path) / 1024:.1f} KB")
    
    return output_path
